import openpyxl

#criar planilha(book)
book = openpyxl.Workbook()
#visualizar paginas existentes
print(book.sheetnames)
#criar uma pagina
book.create_sheet('Frutas')
#selecionar uma pagina
frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])
frutas_page.append(['Banana','5','R$3,90'])
frutas_page.append(['Fruta','2','R$14,20'])
frutas_page.append(['Fruta','3','R$30,30'])
frutas_page.append(['Fruta','4','R$50,10'])
#salvar planilha
book.save('Planilha de Compras.xlsx')
